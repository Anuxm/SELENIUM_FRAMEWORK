import openpyxl
#load excel data
book=openpyxl.load_workbook(r"D:\Book1.xlsx")
#to get control of active sheet object
sheet=book.active
#print first row
# cell=sheet.cell(row=1,column=2)
#exctrac value of cell
# print(cell.value)
# #write 
# k=sheet.cell(row=2,column=2).value
# print(k)
# print(sheet.max_row)
# print(sheet.max_column)
# z=sheet['A5'].value
# print(z)
# for i in range(1,sheet.max_row+1):
#     print(sheet.cell(row=i,column=1).value)

# for i in range(1,sheet.max_row+1):
#     for j in range(1,sheet.max_column+1):
#         print(sheet.cell(row=i,column=j).value)

# for i in range(1,sheet.max_row+1):
#     if sheet.cell(row=i,column=1).value=="test3":
#         for j in range(1,sheet.max_column+1):
#             print(sheet.cell(row=i,column=j).value)
dict={}
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="test3":
        for j in range(2,sheet.max_column+1):
            dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(dict)