import openpyxl
#i dont want to create object for this so i use static method--without creating object we can call
#cls.methodname,self not required

class TestEdata:
    #  test_data=[{"firstname":"ananya","email":"anu37858@gmail.com"},{"firstname":"anu","email":"akoljn@gmail.com"}]
     @staticmethod
     def get_exc_data(test_case_name):
        di={}   
        book=openpyxl.load_workbook(r"D:\Book1.xlsx")
        sheet=book.active
        for i in range(1,sheet.max_row+1):
            if sheet.cell(row=i,column=1).value==test_case_name:
                for j in range(2,sheet.max_column+1):
                    di[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
        return[di]